import matplotlib.pyplot as mat;
import statsmodels.api as stats;
import openpyxl
import numpy as np
wb = openpyxl.load_workbook('copdtrends.xlsx', data_only=True)
print(wb.sheetnames)
copdSheet = wb["COPD&Asthma"]
smokeSheet = wb["Smokers"]
heatSheet = wb['Heat']
wfSheet= wb['California Wildfires']

smokeList = []
for s in range(4, 12):#reads all values as floats
    smokeList.append(smokeSheet.cell(row=s, column=10).value)

def makeCountyHeatCOPDGraph(county):
    copdRow = 0
    heatRow = 0
    for i in range(0,57):
        if heatSheet.cell(row=i+2, column=7).value==county:
            heatRow = i+2; copdRow= i+4; break;
    tempDiffDict={}
    for yr in range(0,8):
        thisYrCopdRate = float(str(copdSheet.cell(row=copdRow,column=yr*2+1).value).split(' ')[2])/100
        tempDiffDict[heatSheet.cell(row=heatRow, column=yr+9).value]= thisYrCopdRate - smokeList[yr]

    xList=tempDiffDict.keys()
    yList=tempDiffDict.values()
    mat.scatter(xList,yList)
    mat.title('Temperature on COPD rates in '+county)
    mat.ylabel('COPD Proportion - Smoke Proportion')
    mat.xlabel('Temperature')
    #adding linear regression
    
    m, b = np.polyfit(list(xList), list(yList), 1)#calc line of best fit
    yArray = []
    for n in xList:
        yArray.append(m*n+b)
    mat.plot(xList, yArray)# plot line
    
    xList = stats.add_constant(list(xList)) #mx + b formatting
    model = stats.OLS(list(yList), list(xList)).fit()
    print(model.summary())# generating p-value
    #print(model.params
    
    mat.show()
    

def makeCountyWFCOPDGraph(county):
    #find in wild fires list
    copdRow = 0
    wfRow = 0
    for i in range(0,58):
        if heatSheet.cell(row=i+2, column=7).value==county:
            copdRow= i+4; break;
    for i in range(0,49):
        thisCounty = str(wfSheet.cell(row=i+1,column=1).value).split(' ')[1]
        if thisCounty==county:
            wfRow= i+1; break;
    wfDiffDict={}
    for yr in range(0,8):
        thisYrAcres = str(wfSheet.cell(row=wfRow, column=yr*2+1).value).split(' ')[2]
        thisYrCopdRate = float(str(copdSheet.cell(row=copdRow,column=yr*2+1).value).split(' ')[2])/100
        wfDiffDict[thisYrAcres]= thisYrCopdRate - smokeList[yr]
    
    xList=np.array(list(wfDiffDict.keys()), dtype=np.float32)
    yList=np.array(list(wfDiffDict.values()), dtype=np.float32)
    mat.scatter(xList,yList)
    mat.title('Wildfires on COPD rates in '+county)
    mat.ylabel('COPD Proportion - Smoke Proportion')
    mat.xlabel('Acres Burned in Wildfires')
    #adding linear regression
    
    m, b = np.polyfit(xList, yList, 1)#calc line of best fit
    yArray = []
    for n in xList:
        yArray.append(m*n+b)
    mat.plot(xList, yArray)# plot line
    
    xList = stats.add_constant(xList) #mx + b formatting
    model = stats.OLS(yList, xList).fit()
    print(model.summary())# generating p-value
    #print(model.params
    
    mat.show()

#user's command   
#makeCountyHeatCOPDGraph('Humboldt')
makeCountyWFCOPDGraph('Yuba')
    