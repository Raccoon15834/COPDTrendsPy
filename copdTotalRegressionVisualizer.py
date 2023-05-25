  
import matplotlib.pyplot as mat;
import statsmodels.api as stats;
import openpyxl
import numpy as np

wb = openpyxl.load_workbook('copdtrends.xlsx', data_only=True)
print(wb.sheetnames)
copdSheet = wb["COPD&Asthma"]
smokeSheet = wb["Smokers"]
heatSheet = wb['Heat']
wfSheet= wb['California Wildfires']

smokeList = []
for s in range(4, 12):#reads all values as floats
    smokeList.append(smokeSheet.cell(row=s, column=10).value)

def summaryStats(xList):
    quartiles = np.percentile(xList, [25, 50, 75])
    print("min "+str(min(xList)))
    print(quartiles[0])
    print(quartiles[1])
    print(quartiles[2])
    print("max "+str(max(xList)))
    print("Standard dev "+str(np.std(xList)))
    print("mean "+str(sum(xList)/len(xList)))

def makeTotalHeatCOPDGraph():
    tempDiffDict={}
    for i in range(0,58):
        copdRow = i+4
        heatRow = i+2
        for yr in range(0,8):
            thisYrCopdRate = float(str(copdSheet.cell(row=copdRow,column=yr*2+1).value).split(' ')[2])/100
            tempDiffDict[heatSheet.cell(row=heatRow, column=yr+9).value]= thisYrCopdRate - smokeList[yr]

    xList=np.array(list(tempDiffDict.keys()), dtype=np.float32)
    yList=np.array(list(tempDiffDict.values()), dtype=np.float32)
    mat.scatter(xList,yList)
    mat.title('Temperature on COPD rates in California\'s Counties')
    mat.ylabel('COPD Proportion - Smoke Proportion')
    mat.xlabel('Temperature')
    
    print("Asthma-COPD summary stats")
    summaryStats(xList)
    #adding linear regression
    m, b = np.polyfit(list(xList), list(yList), 1)
    yArray = []
    for n in xList:
        yArray.append(m*n+b)
    mat.plot(xList, yArray)
    
    xList = stats.add_constant(list(xList)) #mx + b formatting
    model = stats.OLS(list(yList), list(xList)).fit()
    print(model.summary())
    #print(model.params)
    
    mat.show()

def makeTotalWFCOPDGraph():
    wfDiffDict={}
    for i in range(0,47):
        evalulateDict=True
        wfRow=i+1
        copdRow = 0
        county = str(wfSheet.cell(row=i+1,column=1).value).split(' ')[1]
        for i in range(0,58):
            if i==57:
                evalulateDict=False; break;
            if heatSheet.cell(row=i+2, column=7).value in county:
                copdRow= i+4; break;
        for yr in range(0,8):
            if evalulateDict==False: break;
            thisYrAcres = float(str(wfSheet.cell(row=wfRow, column=yr*2+1).value).split(' ')[2])
            if thisYrAcres>1688: break;# get rid of extreme outliers
            thisYrCopdRate = float(str(copdSheet.cell(row=copdRow,column=yr*2+1).value).split(' ')[2])/100
            wfDiffDict[thisYrAcres]= thisYrCopdRate - smokeList[yr]
        
    xList=np.array(list(wfDiffDict.keys()), dtype=np.float32)
    yList=np.array(list(wfDiffDict.values()), dtype=np.float32)
    mat.scatter(xList,yList)
    mat.title("Wildfires on COPD rates in California\'s counties")
    mat.ylabel('COPD Proportion - Smoke Proportion')
    mat.xlabel('Acres Burned in Wildfires')
    #calculating the 5 number summary, mean and stddev, min, max
    print("Wildfire acres summary stats")
    summaryStats(xList)
    #adding linear regression
        
    m, b = np.polyfit(xList, yList, 1)#calc line of best fit
    yArray = []
    for n in xList:
        yArray.append(m*n+b)
    mat.plot(xList, yArray)# plot line
    xList = stats.add_constant(xList) #mx + b formatting
    model = stats.OLS(yList, xList).fit()
    print(model.summary())# generating p-value
    
    mat.show()
    
        
#user's command   
makeTotalHeatCOPDGraph()
#makeTotalWFCOPDGraph()
    